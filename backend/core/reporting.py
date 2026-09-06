import csv
from io import StringIO, BytesIO
from decimal import Decimal
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import serializers
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from core.permissions import IsInternalUser
from core.access import scoped_quotes, quote_for, require_roles
from quotations.models import ApprovalLog


def report_quotes(request):
    qs = scoped_quotes(request.user)
    for key, lookup in [('from', 'created_at__date__gte'), ('to', 'created_at__date__lte')]:
        if request.query_params.get(key):
            value = serializers.DateField().run_validation(request.query_params[key])
            qs = qs.filter(**{lookup: value})
    for key, lookup in [('rep', 'rep_id'), ('status', 'status'), ('category', 'lines__product__category')]:
        if request.query_params.get(key):
            qs = qs.filter(**{lookup: request.query_params[key]})
    return qs.distinct()


def safe_cell(value):
    text = str(value)
    return "'" + text if text.lstrip().startswith(('=', '+', '-', '@', '\t', '\r')) else text


@api_view(['GET'])
@permission_classes([IsInternalUser])
def reports(request):
    quotes = list(report_quotes(request))
    rows = [{'id': q.id, 'quote_number': q.quote_number, 'customer': q.customer.name,
             'rep': q.rep.get_full_name() if q.rep else 'Unassigned', 'status': q.status,
             'amount': q.total_amount, 'discount': q.total_discount, 'margin': q.margin_pct,
             'created': q.created_at.date()} for q in quotes]
    export_fmt = request.query_params.get('export') or request.query_params.get('format')
    if export_fmt in ('xlsx', 'pdf'):
        return report_file(rows, export_fmt)
    if export_fmt == 'csv':
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['Quotation', 'Customer', 'Rep', 'Status', 'Net amount (INR)', 'Discount (INR)', 'Margin %', 'Created'])
        for row in rows:
            writer.writerow([safe_cell(row[k]) for k in ['quote_number','customer','rep','status','amount','discount','margin','created']])
        response = HttpResponse('\ufeff' + output.getvalue(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="dealflow-report.csv"'
        return response
    stages = {}
    for row in rows:
        stages[row['status']] = stages.get(row['status'], 0) + 1
    won = [r for r in rows if r['status'] in ('confirmed', 'fulfillment', 'invoiced', 'paid')]
    closed = len(won) + sum(1 for r in rows if r['status'] in ('rejected', 'cancelled'))
    return Response({'rows': rows, 'stages': stages, 'currency': 'INR', 'summary': {
        'count': len(rows), 'pipeline': sum((r['amount'] for r in rows if r['status'] in ('draft','pending_approval','approved','sent','under_negotiation')), Decimal(0)),
        'bookings': sum((r['amount'] for r in won), Decimal(0)),
        'win_rate': round(len(won) / closed * 100, 1) if closed else 0,
        'discount': sum((r['discount'] for r in rows), Decimal(0)),
        'margin': round(sum(r['margin'] for r in rows) / len(rows), 1) if rows else 0}})


@api_view(['GET'])
@permission_classes([AllowAny])
def quotation_pdf(request, pk):
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.barcode.qr import QrCodeWidget
    from xml.sax.saxutils import escape
    from django.conf import settings
    from core.verification import generate_quotation_signature, amount_to_words_inr, verify_quotation_signature
    from quotations.models import Quotation
    from portal.models import PortalToken

    token_param = request.query_params.get('token')
    sig_param = request.query_params.get('sig')
    
    if request.user and request.user.is_authenticated and getattr(request.user, 'role', '') != 'customer':
        q = quote_for(request, pk)
    elif token_param:
        from portal.views import resolve_token
        q = resolve_token(token_param, pk=pk)
    else:
        q = Quotation.objects.filter(pk=pk).first()
        if not q:
            return Response({'detail': 'Quotation not found.'}, status=404)
        if sig_param and not verify_quotation_signature(q, sig_param):
            return Response({'detail': 'Invalid or tampered document signature.'}, status=403)
        elif not sig_param:
            if not request.user or not request.user.is_authenticated:
                return Response({'detail': 'Authentication or valid verification token required.'}, status=401)
            q = quote_for(request, pk)

    active_token = PortalToken.objects.filter(quotation=q, is_used=False, expires_at__gt=timezone.now()).first()
    token_str = str(active_token.token) if active_token else (q.portal_token or '')
    
    sig = generate_quotation_signature(q)
    origin = request.META.get('HTTP_ORIGIN') or request.META.get('HTTP_REFERER') or ''
    if origin and 'vercel.app' in origin:
        from urllib.parse import urlparse
        p = urlparse(origin)
        frontend_base = f"{p.scheme}://{p.netloc}"
    else:
        frontend_base = getattr(settings, 'FRONTEND_URL', 'https://deal-flow360-omega.vercel.app').rstrip('/')
        if 'dealflow360.vercel.app' in frontend_base:
            frontend_base = 'https://deal-flow360-omega.vercel.app'
    verify_url = f"{frontend_base}/verify/{q.quote_number}?sig={sig}"
    if token_str:
        verify_url += f"&token={token_str}"

    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    styles = getSampleStyleSheet()
    
    c_primary = colors.HexColor('#0F172A')
    c_brand = colors.HexColor('#2563EB')
    c_brand_dark = colors.HexColor('#1E40AF')
    c_surface = colors.HexColor('#F8FAFC')
    c_border = colors.HexColor('#E2E8F0')
    c_muted = colors.HexColor('#64748B')

    style_company_name = ParagraphStyle('CompName', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=18, leading=22, textColor=c_brand_dark)
    style_meta_label = ParagraphStyle('MetaLabel', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=c_muted)
    style_meta_val = ParagraphStyle('MetaVal', parent=styles['Normal'], fontName='Helvetica', fontSize=9, leading=12, textColor=c_primary)
    style_body = ParagraphStyle('BodyCustom', parent=styles['Normal'], fontName='Helvetica', fontSize=8.5, leading=11, textColor=c_primary)
    style_table_header = ParagraphStyle('TH', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=colors.white)
    style_table_cell = ParagraphStyle('TC', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=10, textColor=c_primary)
    style_table_cell_bold = ParagraphStyle('TCB', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=c_primary)
    style_table_cell_right = ParagraphStyle('TCR', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=10, textColor=c_primary, alignment=2)
    style_table_cell_right_bold = ParagraphStyle('TCRB', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=c_primary, alignment=2)
    style_verify_text = ParagraphStyle('VerifyText', parent=styles['Normal'], fontName='Helvetica', fontSize=7, leading=9, textColor=c_muted)

    content = []

    header_left = [
        Paragraph("DEALFLOW360", style_company_name),
        Paragraph("<font color='#2563EB'><b>Enterprise Quote-to-Cash Systems</b></font>", style_body),
        Spacer(1, 4),
        Paragraph("<b>GSTIN:</b> 27AAACD8921M1Z4 · <b>PAN:</b> AAACD8921M", style_verify_text),
        Paragraph("Tower 4, World Trade Centre, Pune, MH 411014 · support@dealflow360.com", style_verify_text),
    ]

    status_color = '#059669' if q.status in ('approved', 'confirmed', 'paid') else ('#D97706' if q.status in ('pending_approval', 'under_negotiation') else '#2563EB')
    header_right = [
        Paragraph(f"<font size='14'><b>TAX QUOTATION</b></font>", ParagraphStyle('RHead', parent=styles['Normal'], alignment=2, textColor=c_brand_dark)),
        Paragraph(f"<b>Quote No:</b> <font color='#2563EB'>{escape(q.quote_number)}</font>", ParagraphStyle('RNum', parent=styles['Normal'], alignment=2, fontSize=10, leading=13)),
        Paragraph(f"<b>Date:</b> {q.created_at.strftime('%d %b %Y')} · <b>Valid Until:</b> {q.valid_until.strftime('%d %b %Y') if q.valid_until else '30 Days'}", ParagraphStyle('RDate', parent=styles['Normal'], alignment=2, fontSize=8, leading=11, textColor=c_muted)),
        Paragraph(f"<b>Status:</b> <font color='{status_color}'><b>{q.get_status_display().upper()}</b></font>", ParagraphStyle('RStat', parent=styles['Normal'], alignment=2, fontSize=8.5, leading=11)),
    ]

    header_table = Table([[header_left, header_right]], colWidths=[310, 213])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 0),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    content.append(header_table)
    content.append(Spacer(1, 8))
    content.append(HRFlowable(width="100%", thickness=1.5, color=c_brand, spaceBefore=2, spaceAfter=8))

    cust = q.customer
    customer_info = [
        Paragraph("<b>BILLED & SHIPPED TO:</b>", style_meta_label),
        Spacer(1, 2),
        Paragraph(f"<b>{escape(cust.name)}</b>", style_meta_val),
        Paragraph(escape(cust.company or 'Direct Account'), style_body),
        Paragraph(escape(cust.address or 'Standard Shipping Address'), style_verify_text),
        Paragraph(f"<b>Email:</b> {escape(cust.email)} | <b>Phone:</b> {escape(cust.phone or 'N/A')}", style_verify_text),
        Paragraph(f"<b>Client Tier:</b> {cust.get_tier_display()}", style_verify_text),
    ]

    rep_info = [
        Paragraph("<b>COMMERCIAL TERMS & REPRESENTATIVE:</b>", style_meta_label),
        Spacer(1, 2),
        Paragraph(f"<b>Sales Rep:</b> {escape(q.rep.get_full_name() if q.rep else 'Enterprise Sales Desk')}", style_body),
        Paragraph(f"<b>Payment Terms:</b> {escape(q.payment_terms)}", style_body),
        Paragraph(f"<b>Currency:</b> Indian Rupee (INR · ₹)", style_body),
        Paragraph(f"<b>Approval Level:</b> {q.get_required_approval_level_display()}", style_verify_text),
    ]

    party_table = Table([[customer_info, rep_info]], colWidths=[270, 253])
    party_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), c_surface),
        ('BOX', (0,0), (-1,-1), 0.5, c_border),
        ('TOPPADDING', (0,0), (-1,-1), 7),
        ('BOTTOMPADDING', (0,0), (-1,-1), 7),
        ('LEFTPADDING', (0,0), (-1,-1), 9),
        ('RIGHTPADDING', (0,0), (-1,-1), 9),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    content.append(party_table)
    content.append(Spacer(1, 10))

    table_data = [[
        Paragraph("<b>#</b>", style_table_header),
        Paragraph("<b>Item & Description</b>", style_table_header),
        Paragraph("<b>Qty</b>", style_table_header),
        Paragraph("<b>Rate (₹)</b>", ParagraphStyle('THRight', parent=style_table_header, alignment=2)),
        Paragraph("<b>Disc %</b>", ParagraphStyle('THRight2', parent=style_table_header, alignment=2)),
        Paragraph("<b>GST %</b>", ParagraphStyle('THRight3', parent=style_table_header, alignment=2)),
        Paragraph("<b>Net Total (₹)</b>", ParagraphStyle('THRight4', parent=style_table_header, alignment=2)),
    ]]

    for idx, line in enumerate(q.lines.select_related('product', 'variant'), 1):
        item_title = f"<b>{escape(line.product.name)}</b>"
        if line.variant:
            item_title += f" <font color='#64748B'>({escape(line.variant.attribute)}: {escape(line.variant.value)})</font>"
        
        desc_parts = []
        if line.product.sku:
            desc_parts.append(f"SKU: {escape(line.product.sku)}")
        if line.is_subscription:
            desc_parts.append("<font color='#2563EB'>[Recurring Subscription]</font>")
        if line.description:
            desc_parts.append(escape(line.description))
        
        desc_text = " · ".join(desc_parts)
        cell_item = [
            Paragraph(item_title, style_table_cell),
            Paragraph(f"<font size='6.5' color='#64748B'>{desc_text}</font>", style_table_cell) if desc_text else Spacer(1, 0)
        ]

        tax_rate = line.tax_pct if line.tax_pct is not None else line.product.tax_pct
        table_data.append([
            Paragraph(str(idx), style_table_cell),
            cell_item,
            Paragraph(f"{line.qty:g}", style_table_cell),
            Paragraph(f"{line.unit_price:,.2f}", style_table_cell_right),
            Paragraph(f"{line.discount_pct:g}%", style_table_cell_right),
            Paragraph(f"{tax_rate:g}%", style_table_cell_right),
            Paragraph(f"{line.line_total:,.2f}", style_table_cell_right_bold),
        ])

    items_table = Table(table_data, colWidths=[18, 215, 32, 75, 45, 45, 93], repeatRows=1)
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), c_brand_dark),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (0,0), (-1,-1), 5),
        ('RIGHTPADDING', (0,0), (-1,-1), 5),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, c_border),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#FAFBFD')]),
    ]))
    content.append(items_table)
    content.append(Spacer(1, 8))

    gross_total = q.gross_total
    discount_total = q.total_discount
    net_subtotal = q.total_amount
    tax_total = q.tax_amount
    grand_total = net_subtotal + tax_total
    amount_words = amount_to_words_inr(grand_total)

    summary_left = [
        Paragraph("<b>AMOUNT IN WORDS:</b>", style_meta_label),
        Paragraph(f"<i>{amount_words}</i>", ParagraphStyle('AmtWords', parent=styles['Normal'], fontName='Helvetica-Oblique', fontSize=7.5, leading=9.5, textColor=c_primary)),
        Spacer(1, 6),
        Paragraph("<b>BANK & SETTLEMENT DETAILS:</b>", style_meta_label),
        Paragraph("Bank: HDFC Bank Ltd. · A/C: 50200088991122 · IFSC: HDFC0000123", style_verify_text),
        Paragraph("Payment Reference: <b>" + escape(q.quote_number) + "</b>", style_verify_text),
        Paragraph(f"Notes: {escape(q.notes or 'Subject to standard enterprise warranty and SLA terms.')}", style_verify_text),
    ]

    summary_rows = [
        [Paragraph("Gross Subtotal:", style_table_cell), Paragraph(f"₹ {gross_total:,.2f}", style_table_cell_right)],
        [Paragraph("Volume / Tier Discount:", style_table_cell), Paragraph(f"- ₹ {discount_total:,.2f}", style_table_cell_right)],
        [Paragraph("<b>Taxable Net Amount:</b>", style_table_cell_bold), Paragraph(f"<b>₹ {net_subtotal:,.2f}</b>", style_table_cell_right_bold)],
        [Paragraph("GST Output Tax:", style_table_cell), Paragraph(f"₹ {tax_total:,.2f}", style_table_cell_right)],
        [Paragraph("<font size='8.5'><b>GRAND TOTAL (INR):</b></font>", ParagraphStyle('GT1', parent=style_table_cell_bold, textColor=c_brand_dark)),
         Paragraph(f"<font size='8.5'><b>₹ {grand_total:,.2f}</b></font>", ParagraphStyle('GT2', parent=style_table_cell_right_bold, textColor=c_brand_dark))],
    ]
    summary_right_table = Table(summary_rows, colWidths=[120, 100])
    summary_right_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 2.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2.5),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ('LINEBELOW', (0,2), (-1,2), 0.5, c_border),
        ('BACKGROUND', (0,4), (-1,4), colors.HexColor('#EFF6FF')),
        ('BOX', (0,4), (-1,4), 1, c_brand),
    ]))

    calc_table = Table([[summary_left, summary_right_table]], colWidths=[293, 230])
    calc_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    content.append(calc_table)
    content.append(Spacer(1, 10))

    qr_widget = QrCodeWidget(verify_url)
    qr_widget.barWidth = 64
    qr_widget.barHeight = 64
    qr_widget.qrVersion = 3
    qr_drawing = Drawing(64, 64)
    qr_drawing.add(qr_widget)

    security_cert_text = [
        Paragraph("<font color='#059669'><b>🛡️ CRYPTOGRAPHIC VERIFICATION SEAL</b></font>", ParagraphStyle('CertH', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, leading=10)),
        Spacer(1, 2),
        Paragraph(f"<b>HMAC-SHA256 Hash:</b> <font face='Courier' size='6.5'>{sig}</font>", style_verify_text),
        Paragraph(f"<b>Live Verification Node:</b> <font color='#2563EB'>{escape(verify_url[:70])}...</font>", style_verify_text),
        Paragraph("Scan QR code with any smartphone to verify authenticity directly against DealFlow360 immutable database.", style_verify_text),
        Paragraph("<b>Authorized Digital Seal:</b> DealFlow360 Automated Governance Engine · " + timezone.now().strftime('%Y-%m-%d %H:%M:%S UTC'), style_verify_text),
    ]

    sec_table = Table([[qr_drawing, security_cert_text]], colWidths=[72, 451])
    sec_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F0FDF4')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#86EFAC')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (0,0), (-1,-1), 7),
        ('RIGHTPADDING', (0,0), (-1,-1), 7),
    ]))
    content.append(KeepTogether(sec_table))

    def page_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(c_muted)
        canvas.drawString(36, 18, "DealFlow360 Enterprise Deal Engine · Computer Generated Quotation · No Physical Signature Required")
        canvas.drawRightString(559, 18, f"Page {doc.page} of 1 · Verified ID: {q.quote_number}")
        canvas.restoreState()

    doc.build(content, onFirstPage=page_footer, onLaterPages=page_footer)
    
    response = HttpResponse(output.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{q.quote_number}.pdf"'
    response['X-Verification-Hash'] = sig
    return response


@api_view(['POST'])
@permission_classes([IsInternalUser])
def nudge(request, pk):
    require_roles(request.user, 'sales_manager', 'finance', 'admin')
    q = quote_for(request, pk, lock=True)
    note = serializers.CharField(max_length=1000).run_validation(request.data.get('reason', 'Manager requested an update on this deal.'))
    ApprovalLog.objects.create(quotation=q, actor=request.user, action='escalated', note=note, role_required='sales_rep')
    return Response({'message': 'Escalation recorded in the deal activity trail.'})



def report_file(rows, kind):
    keys = ['quote_number','customer','rep','status','amount','discount','margin','created']
    headers = ['Quotation','Customer','Representative','Stage','Net INR','Discount INR','Margin %','Created']
    output = BytesIO()
    if kind == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Deal performance'
        sheet.append(headers)
        for row in rows:
            sheet.append([safe_cell(row[k]) if isinstance(row[k],str) else row[k] for k in keys])
        for cell in sheet[1]:
            cell.fill = PatternFill('solid',fgColor='245B49')
            cell.font = Font(color='FFFFFF',bold=True)
            cell.alignment = Alignment(vertical='center')
        sheet.row_dimensions[1].height = 28
        for index,width in enumerate([23,34,24,24,20,20,15,18],1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='center')
            for index in (4,5):
                row[index].number_format = '"INR "#,##0.00'
            row[7].number_format = 'dd mmm yyyy'
        sheet.freeze_panes='A2'
        sheet.auto_filter.ref=sheet.dimensions
        workbook.save(output)
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    else:
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from xml.sax.saxutils import escape
        styles=getSampleStyleSheet()
        styles['Normal'].fontSize=8
        styles['Normal'].leading=12
        doc=SimpleDocTemplate(output,pagesize=landscape(A4),leftMargin=32,rightMargin=32,topMargin=32,bottomMargin=32)
        content=[Paragraph('DealFlow360 | Deal performance',styles['Title']),
            Paragraph(f'{len(rows)} quotations | Currency: INR | Generated {timezone.localdate()}',styles['Normal']),Spacer(1,20)]
        data=[headers]
        for row in rows:
            data.append([Paragraph(escape(f'{row[k]:,.2f}' if k in ('amount','discount','margin') else str(row[k])),styles['Normal']) for k in keys])
        table=Table(data,colWidths=[90,150,95,90,100,95,60,85],repeatRows=1)
        table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#245B49')),('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTSIZE',(0,0),(-1,-1),8),('TOPPADDING',(0,0),(-1,-1),10),('BOTTOMPADDING',(0,0),(-1,-1),10),
            ('VALIGN',(0,0),(-1,-1),'TOP'),('LINEBELOW',(0,0),(-1,-1),.5,colors.HexColor('#E1E6D9'))]))
        content.append(table)
        def footer(canvas, doc):
            canvas.setFont('Helvetica',8)
            canvas.setFillColor(colors.HexColor('#6D7B60'))
            canvas.drawString(32,18,'DealFlow360 - Internal sales report')
            canvas.drawRightString(810,18,f'Page {doc.page}')
        doc.build(content,onFirstPage=footer,onLaterPages=footer)
        content_type='application/pdf'
    response=HttpResponse(output.getvalue(),content_type=content_type)
    response['Content-Disposition']=f'attachment; filename="dealflow-report.{kind}"'
    return response
